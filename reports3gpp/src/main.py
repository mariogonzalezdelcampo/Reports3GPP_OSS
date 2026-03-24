"""
Main entry point for the 3GPP report generator.
"""
import os
import sys
import logging
import re
import shutil
from pathlib import Path
from typing import Optional

# Ensure the src directory is on the import path
sys.path.insert(0, str(Path(__file__).parent))

from config import AppConfig
from downloader import download_file, download_zip_ftp, download_zip
from extractor import extract_zip
from excel_processor import excel_to_csv, filter_items
from tdoc_handler import process_tdoc
from summary import append_summary
from html_parser import find_zip_file_in_html
from ollama_client import query_ollama
from docx import Document

# ---------------------------------------------------------------------------
# Helper to download and extract the "Charging exec report" ZIP referenced in the
# Excel file. This is used both in BYPASS and FULL processing modes.
# ---------------------------------------------------------------------------
def _extract_charging_report(excel_file: Path, cfg) -> None:
    """Locate the row with Title "Charging exec report" and extract its ZIP.

    The function reads the Excel file with ``openpyxl`` (read‑only mode) to find
    the column indices for ``Title`` and ``TDoc`` (case‑insensitive). When the
    matching row is found, the URL from the ``TDoc`` column is downloaded –
    supporting both FTP and HTTP – and extracted into the ``summaries``
    directory of the meeting. Errors are logged but do not raise, so the main
    pipeline can continue.
    """
    try:
        from openpyxl import load_workbook

        # Load workbook in normal mode to access hyperlinks.
        wb = load_workbook(excel_file, read_only=False, data_only=True)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        title_idx = None
        tdoc_idx = None
        for i, col_name in enumerate(header):
            if isinstance(col_name, str):
                lowered = col_name.lower()
                if lowered == "title":
                    title_idx = i
                elif lowered == "tdoc":
                    tdoc_idx = i
        if title_idx is None or tdoc_idx is None:
            logger.warning("Title or TDoc column not found in Excel; skipping charging report extraction")
            wb.close()
            return

        # The summaries directory should be inside the meeting folder, not one level up.
        # ``excel_file.parent`` is the meeting directory (e.g., ``.../3GPP_meeting_docs_165``).
        summaries_dir = excel_file.parent / "summaries"
        summaries_dir.mkdir(exist_ok=True)

        for row in ws.iter_rows(min_row=2):
            title_cell = row[title_idx]
            if title_cell.value and str(title_cell.value).strip() == "Charging exec report":
                tdoc_cell = row[tdoc_idx]
                zip_url = None
                if tdoc_cell.hyperlink:
                    zip_url = tdoc_cell.hyperlink.target
                elif tdoc_cell.value:
                    zip_url = str(tdoc_cell.value).strip()
                if not zip_url:
                    logger.warning("Charging exec report row has no TDoc URL")
                    break
                # Ensure temporary directory exists.
                cfg.temp_dir.mkdir(parents=True, exist_ok=True)
                if zip_url.startswith("ftp://"):
                    from urllib.parse import urlparse
                    parsed = urlparse(zip_url)
                    host = parsed.hostname
                    path = parsed.path
                    temp_zip = cfg.temp_dir / f"temp_{os.path.basename(path)}"
                    download_zip_ftp(host, path, temp_zip)
                    extract_zip(temp_zip, summaries_dir)
                    temp_zip.unlink()
                else:
                    zip_path = download_zip(zip_url, cfg.temp_dir)
                    extract_zip(zip_path, summaries_dir)
                    zip_path.unlink()
                logger.info("Extracted Charging exec report into %s", summaries_dir)
                break
        wb.close()
    except Exception as e:
        logger.error("Failed to extract Charging exec report: %s", e)

# ---------------------------------------------------------------------------
# Logging configuration
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    handlers=[
        logging.FileHandler("app.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helper to persist LLM interactions
# ---------------------------------------------------------------------------
def _log_interaction(prompt: str, response: str, *, mode: str = "w") -> None:
    """Write the *prompt* and *response* to ``REQUEST.md`` and ``RESPONSE.md``.

    A separator line (``---``) is added after each entry so that multiple
    interactions are clearly delimited in both files. The ``mode`` argument
    determines whether the files are overwritten (``"w"``) or appended to
    (``"a"``).
    """
    request_path = Path("REQUEST.md")
    response_path = Path("RESPONSE.md")
    # Write the prompt and a separator
    with open(request_path, mode, encoding="utf-8") as f_req:
        f_req.write(prompt + "\n")
        f_req.write("---\n")
    # Write the response and a separator
    with open(response_path, mode, encoding="utf-8") as f_res:
        f_res.write(response + "\n")
        f_res.write("---\n")

# ---------------------------------------------------------------------------
# Text extraction / cleaning utilities for DOCX files
# ---------------------------------------------------------------------------
_IGNORE_HEADINGS = {
    "foreword",
    "scope",
    "references",
    "definition of terms",
    "change history",
    "versioning",
    "modal-verb conventions",
    "definitions",
    "symbols",
    "abbreviations",
}

def _clean_text(text: str) -> str:
    """Normalise whitespace and remove non‑printable characters."""
    text = re.sub(r"[\x00-\x1F\x7F]", " ", text)
    text = re.sub(r" +", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()

def _extract_docx_text(docx_path: Path) -> str:
    """Extract text from *docx_path* while skipping irrelevant headings."""
    doc = Document(str(docx_path))
    lines = []
    for para in doc.paragraphs:
        txt = para.text.strip()
        if not txt:
            continue
        if txt.lower() in _IGNORE_HEADINGS:
            continue
        lines.append(txt)
    return "\n".join(lines)

# ---------------------------------------------------------------------------
# Helper to summarise all DOCX files in a folder and write a single .md file.
# ---------------------------------------------------------------------------
def _summarise_folder(folder_path: Path, cfg) -> None:
    """Summarise every DOCX in *folder_path* and write a .md named after the folder.

    The output file ``<folder_name>.md`` will contain, for each DOCX, the file name
    followed by its summary. The prompt used is taken from ``cfg.fixed_prompt``.
    """
    docx_files = sorted(folder_path.glob("*.docx"))
    if not docx_files:
        logger.info("No DOCX files found in %s", folder_path)
        return

    summary_path = folder_path / f"{folder_path.name}.md"
    with open(summary_path, "w", encoding="utf-8") as summary_file:
            for docx_file in docx_files:
                logger.info(f"Processing DOCX file: {docx_file.name}")
                raw_text = _extract_docx_text(docx_file)
                cleaned = _clean_text(raw_text)
                max_chunk = 200000
                chunks = [cleaned[i : i + max_chunk] for i in range(0, len(cleaned), max_chunk)]
                fixed_prompt = cfg.fixed_prompt or ""
                summary_parts = []
                for chunk in chunks:
                    prompt_chunk = fixed_prompt + chunk + "\n=== FIN DOCUMENTO ==="
                    chunk_resp = query_ollama(
                        "http://10.95.118.26:11434",
                        "gpt-oss:120b",
                        prompt_chunk,
                        temperature=0.1,
                    )
                    summary_parts.append(chunk_resp.strip())
                # Write filename as a highlighted markdown heading and its summary
                # Using a level‑2 heading with bold for clear separation
                summary_file.write(f"## **{docx_file.name}**\n\n")
                summary_file.write("\n\n".join(summary_parts))
                summary_file.write("\n\n")
    logger.info(f"Summary written to {summary_path}")

# ---------------------------------------------------------------------------
# Main workflow
# ---------------------------------------------------------------------------
def main(meeting_number: Optional[int] = None) -> None:
    cfg = AppConfig.load()
    if meeting_number is not None:
        if meeting_number <= 0:
            raise ValueError("Meeting number must be a positive integer")
        cfg.meeting_number = meeting_number

    mode = cfg.processing_mode.upper() if hasattr(cfg, "processing_mode") else "FULL"

    # -----------------------------------------------------------------------
    # BYPASS mode – only LLM calls, no FTP / Excel processing
    # -----------------------------------------------------------------------
    if mode == "BYPASS":
        logger.info("Processing mode BYPASS – skipping pipeline, invoking LLM only")
        # Simple test prompt
        prompt = "Hola, que tal estas?"
        logger.info(f"Prompt: {prompt}")
        response = query_ollama("http://10.95.118.26:11434", "gpt-oss:120b", prompt)
        logger.info(f"Response: {response}")

        # ---------------------------------------------------------------
        # Summarise all DOCX files in each sub‑folder of the meeting directory
        # ---------------------------------------------------------------
        meeting_dir = cfg.documents_dir / f"3GPP_meeting_docs_{cfg.meeting_number}"
        # Use the module‑level helper defined earlier (requires cfg argument)
        for sub_folder in meeting_dir.iterdir():
            if sub_folder.is_dir():
                _summarise_folder(sub_folder, cfg)

        # After all summaries are generated, move the .md files to a top‑level
        # ``summaries`` directory (created if it does not exist).
        summaries_dir = meeting_dir / "summaries"
        summaries_dir.mkdir(exist_ok=True)
        for sub_folder in meeting_dir.iterdir():
            if sub_folder.is_dir():
                md_file = sub_folder / f"{sub_folder.name}.md"
                if md_file.is_file():
                    shutil.move(str(md_file), str(summaries_dir / md_file.name))

        # -----------------------------------------------------------------
        # In BYPASS mode we still want to extract the Charging exec report.
        # Locate the Excel file (if any) inside the meeting directory and
        # invoke the helper. If the file is missing we simply skip extraction.
        # -----------------------------------------------------------------
        excel_file = next(meeting_dir.rglob("*.xls*"), None)
        if excel_file and excel_file.is_file():
            _extract_charging_report(excel_file, cfg)
        else:
            logger.info("No Excel file found in BYPASS mode – skipping Charging exec report extraction")
        return

    # -----------------------------------------------------------------------
    # FULL processing – download ZIP, extract Excel, filter items, etc.
    # -----------------------------------------------------------------------
    logger.info("Starting report generation for meeting %s", cfg.meeting_number)

    # Build remote FTP path
    host = "ftp.3gpp.org"
    base_path = "/tsg_sa/WG5_TM/TSGS5_"
    remote_path = f"{base_path}{cfg.meeting_number}/Tdoclist/"

    # Locate the ZIP file on the FTP server
    try:
        from downloader import list_ftp_directory
        files = list_ftp_directory(host, remote_path)
        zip_files = [f for f in files if f.endswith('.zip') and 'TDoc_List_Meeting' in f]
        if zip_files:
            zip_name = zip_files[0]
        else:
            zip_files = [f for f in files if f.endswith('.zip')]
            if zip_files:
                zip_name = zip_files[0]
            else:
                raise RuntimeError(f"No ZIP files found in FTP directory {remote_path}")
    except Exception as e:
        raise RuntimeError(f"Failed to find ZIP file in FTP directory {remote_path}: {e}")

    logger.info("Found ZIP file: %s", zip_name)

    # Download and extract the ZIP
    zip_url = f"{base_path}{cfg.meeting_number}/Tdoclist/{zip_name}"
    zip_path = cfg.documents_dir / zip_name
    download_zip_ftp(host, zip_url, zip_path)
    meeting_dir = cfg.documents_dir / f"3GPP_meeting_docs_{cfg.meeting_number}"
    extract_zip(zip_path, meeting_dir)

    # Locate the Excel file and convert it to CSV
    excel_file = next(meeting_dir.rglob("*.xls*"), None)
    if not excel_file:
        raise FileNotFoundError("No Excel file found in the extracted directory")
    csv_file = excel_file.with_suffix('.csv')
    excel_to_csv(excel_file, csv_file)

    # Filter the CSV rows according to the original business rules
    items = filter_items(csv_file)

    # Process each TDoc item
    for i, item in enumerate(items):
        try:
            doc_name = process_tdoc(item, meeting_dir, cfg.temp_dir)
            append_summary(
                meeting_dir / item["Related WIs"],
                item["Related WIs"],
                doc_name,
                item["Spec"],
            )
        except Exception as e:
            logger.error("Failed to process item %d: %s", i, e)
            continue

    logger.info("Report generation completed successfully")

    # After the full pipeline, summarise DOCX files in each sub‑folder of the meeting directory.
    # The meeting_dir variable already points to the extracted meeting folder.
    for sub_folder in meeting_dir.iterdir():
        if sub_folder.is_dir():
            _summarise_folder(sub_folder, cfg)

    # Move generated markdown summaries to a top‑level ``summaries`` folder.
    summaries_dir = meeting_dir / "summaries"
    summaries_dir.mkdir(exist_ok=True)
    for sub_folder in meeting_dir.iterdir():
        if sub_folder.is_dir():
            md_file = sub_folder / f"{sub_folder.name}.md"
            if md_file.is_file():
                shutil.move(str(md_file), str(summaries_dir / md_file.name))

        # Extract the Charging exec report after all other processing.
        _extract_charging_report(excel_file, cfg)
    # -----------------------------------------------------------------
    # Additional step: download the ZIP referenced by the Excel row whose
    # ``Title`` column equals "Charging exec report" and extract its contents
    # (expected to be a .ppt or .pptx) into the ``summaries`` directory.
    # -----------------------------------------------------------------
    try:
        # ``excel_file`` was defined earlier in the FULL pipeline section.
        # Re‑use it here; if it does not exist (e.g., BYPASS mode) we simply
        # skip this step.
        if excel_file and excel_file.is_file():
            from openpyxl import load_workbook

            wb = load_workbook(excel_file, read_only=False, data_only=True)
            ws = wb.active
            # Identify column indices for Title and TDoc (case‑insensitive).
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            title_idx = None
            tdoc_idx = None
            for i, col_name in enumerate(header):
                if isinstance(col_name, str):
                    lowered = col_name.lower()
                    if lowered == "title":
                        title_idx = i
                    elif lowered == "tdoc":
                        tdoc_idx = i
            if title_idx is not None and tdoc_idx is not None:
                for row in ws.iter_rows(min_row=2):
                    title_cell = row[title_idx]
                    if title_cell.value and str(title_cell.value).strip() == "Charging exec report":
                        tdoc_cell = row[tdoc_idx]
                        # Prefer hyperlink target if present, otherwise the cell value.
                        zip_url = None
                        if tdoc_cell.hyperlink:
                            zip_url = tdoc_cell.hyperlink.target
                        elif tdoc_cell.value:
                            zip_url = str(tdoc_cell.value).strip()
                        if zip_url:
                            # Ensure temporary directory exists.
                            cfg.temp_dir.mkdir(parents=True, exist_ok=True)
                            # Determine download method based on scheme.
                            if zip_url.startswith("ftp://"):
                                from urllib.parse import urlparse
                                parsed = urlparse(zip_url)
                                host = parsed.hostname
                                path = parsed.path
                                temp_zip = cfg.temp_dir / f"temp_{os.path.basename(path)}"
                                download_zip_ftp(host, path, temp_zip)
                                extract_zip(temp_zip, summaries_dir)
                                temp_zip.unlink()
                            else:
                                zip_path = download_zip(zip_url, cfg.temp_dir)
                                extract_zip(zip_path, summaries_dir)
                                zip_path.unlink()
                            logger.info("Extracted Charging exec report into %s", summaries_dir)
                        break  # Found the row, no need to continue.
            wb.close()
    except Exception as e:
        # Log the error but do not fail the whole pipeline.
        logger.error("Failed to download or extract Charging exec report: %s", e)

    # -----------------------------------------------------------------------
    # OPTIONAL: final LLM call after the full pipeline (mirrors BYPASS behaviour)
    # -----------------------------------------------------------------------
    # Final LLM call after full processing removed per user request (no REQUEST/RESPONSE files needed)

if __name__ == "__main__":
    meeting_number = None
    if len(sys.argv) > 1:
        try:
            meeting_number = int(sys.argv[1])
        except ValueError:
            logger.error("Invalid meeting number provided")
            sys.exit(1)
    main(meeting_number)