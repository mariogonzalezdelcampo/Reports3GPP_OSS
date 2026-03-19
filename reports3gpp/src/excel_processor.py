"""Utilities for converting Excel files to CSV and filtering data.

The original implementation relied on ``pandas`` which loads the entire workbook
into memory. For large Excel files this caused high memory consumption. The new
implementation streams the workbook using ``openpyxl`` in *read‑only* mode and
writes rows directly to a CSV file with the standard ``csv`` module. It also
preserves the hyperlink URL in the ``TDoc`` column.
"""

import csv
from pathlib import Path
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)


def _load_hyperlink_map(excel_path: Path, tdoc_idx: int) -> Dict[int, str]:
    """Return a mapping of row number → hyperlink target for the TDoc column.

    ``openpyxl`` in ``read_only`` mode does not expose ``cell.hyperlink``. To
    retrieve the URLs we open the workbook a second time in normal mode (still
    ``data_only=True`` to get evaluated values) and iterate over the rows.
    """
    from openpyxl import load_workbook

    hyperlink_map: Dict[int, str] = {}
    wb = load_workbook(excel_path, read_only=False, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):  # skip header
        cell = row[tdoc_idx]
        if cell.hyperlink:
            # ``row[0].row`` gives the 1‑based Excel row number.
            hyperlink_map[cell.row] = cell.hyperlink.target
    wb.close()
    return hyperlink_map


def excel_to_csv(excel_path: Path, csv_path: Path) -> None:
    """Convert an Excel file to CSV while preserving ``TDoc`` hyperlinks.

    The conversion is performed row‑by‑row to keep memory usage low.
    """
    logger.info("Converting %s to %s (streaming, preserving TDoc hyperlinks)", excel_path, csv_path)

    try:
        from openpyxl import load_workbook

        # Load workbook in read‑only mode for streaming.
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        # Extract header row (first row).
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = list(header_row)
        tdoc_idx = None
        if "TDoc" in headers:
            tdoc_idx = headers.index("TDoc")

        # If we need hyperlink values, build a map from row number to URL.
        hyperlink_map: Dict[int, str] = {}
        if tdoc_idx is not None:
            hyperlink_map = _load_hyperlink_map(excel_path, tdoc_idx)

        # Open CSV for writing.
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)  # write header

            # Iterate over data rows while keeping the Excel row number.
            excel_row_num = 2  # first data row in the sheet
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_vals = list(row)
                # Replace the displayed TDoc value with the hyperlink URL if present.
                if tdoc_idx is not None and excel_row_num in hyperlink_map:
                    row_vals[tdoc_idx] = hyperlink_map[excel_row_num]
                writer.writerow(row_vals)
                excel_row_num += 1

        wb.close()
        logger.info("Successfully converted %s to %s", excel_path, csv_path)
    except Exception as e:
        raise RuntimeError(f"Failed to convert Excel to CSV: {e}") from e


def filter_items(csv_path: Path) -> List[Dict]:
    """Filter items from the CSV based on the specified criteria.

    This implementation streams the CSV using ``csv.DictReader`` to keep memory
    usage low.
    """
    logger.info("Filtering items from %s", csv_path)

    required_columns = {"Agenda item", "TDoc Status", "Related WIs", "Spec", "TDoc"}
    allowed_statuses = {"agreed", "approved", "available", "merged"}
    results: List[Dict] = []

    try:
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            # Validate required columns exist
            missing = required_columns.difference(reader.fieldnames or [])
            if missing:
                raise ValueError(f"Missing required columns in CSV: {missing}")

            for row in reader:
                agenda = row.get("Agenda item", "")
                status = row.get("TDoc Status", "").lower()
                if not agenda.startswith("7"):
                    continue
                if status not in allowed_statuses:
                    continue
                results.append(row)

        logger.info("Filtered %d items from %s", len(results), csv_path)
        return results
    except Exception as e:
        raise RuntimeError(f"Failed to filter items from CSV: {e}") from e